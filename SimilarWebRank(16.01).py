import requests
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from xlsxwriter import Workbook
import xlrd
import re
from decimal import *
from openpyxl import load_workbook

workbook = xlrd.open_workbook('binary_result.xlsx')
worksheet = workbook.sheet_by_index(0)
rows = worksheet.nrows
company_links = []
for i in range (rows-1):
     company_links.append(worksheet.cell_value(1+i,3))


ua = UserAgent()
header = {'user-agent': ua.chrome}
for i in range (len(company_links)):
    ua = UserAgent()
    header = {'user-agent': ua.random}
    page = requests.get(
        'https://www.similarweb.com/website/'+company_links[i],
        headers=header)
    soup = BeautifulSoup(page.text, 'lxml')
    b = len(soup.find_all('span',{'class':'rankingItem-value js-countable'}))
    wb = load_workbook(filename='binary_result.xlsx')
    sheet = wb.active
    if b > 0 :
        rating = soup.find('span',{'class':'rankingItem-value js-countable'}).getText()
        temp = int(rating.replace(',',''))
        print(temp)
        if (temp < 1500000):
            print(company_links[i],'rating<1.5m')
            sheet.cell(row=2 + i, column=1).value = temp
        else:
            print(company_links[i],'Low Rating')
            sheet.cell(row=2 + i, column=1).value = temp
    else:           # Проверка, если нет global rank, проверяем по total visits
            c = len(soup.find_all('span',  {'class' :'engagementInfo-valueNumber js-countValue'}))
            if c > 0:
                engage_rating = soup.find('span', {'class' :'engagementInfo-valueNumber js-countValue'}).getText()
                temp2 = int(engage_rating.replace('.','').replace('K','00'))
                print(engage_rating)
                print(temp2)
                if temp2 > 8000:
                    temp3 = str(temp2)+' engage'
                    print(temp3)
                    sheet.cell(row=2 + i, column=1).value = temp3
                else:
                    temp3 = str(temp2) + 'Low Engage'
                    print(company_links[i], temp3)
                    sheet.cell(row=2 + i, column=1).value = 'Low Engage' +'(' + str(temp2) +')'

            else:
                print(company_links[i], 'No rating')
                sheet.cell(row=2 + i, column=1).value = 'No rating'

    wb.save('binary_result.xlsx')


# temp = soup.find_all('div', {'class': 'rankingItem fadeInDown rankingItem--global'})
# newsoup = BeautifulSoup(soup.find_all('div', {'class': 'rankingItem fadeInDown rankingItem--global'}), 'lxml')
# print(newsoup.find_all('span' , {'class':'rankingItem-value js-countable'}))

# f = open('out.html', 'w')
# f.write(soup.prettify())
# f.close()


#
# if len(soup.find_all('code')) == 0:
#     print('!!!!!!!')
#     f = open('out.html', 'w')
#     f.write(soup.prettify())
#     f.close()

        # else:
        #     break










# print(soup.find_all('code')[1:2])
# # print(soup.code.string)
# newsoup = BeautifulSoup(str(soup.find_all('code')), 'lxml')
#
# print(newsoup.get_text())



