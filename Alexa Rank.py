import requests
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from xlsxwriter import Workbook
import xlrd
import re
from decimal import *
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
from selenium.common.exceptions import NoSuchElementException
from random import randint

driver = webdriver.Chrome('C:\\Users\\Admin\\Downloads\\chromedriver')
file_name = 'gambling_11-50.xlsx'
driver.implicitly_wait(4)

def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

company_links = []
wb = load_workbook(filename=file_name)
sheet = wb.active
for i in range (sheet.max_row -1 ):
    company_links.append(sheet.cell(row=2 + i, column=4).value)


for i in range (150,len(company_links)):
    wb = load_workbook(filename=file_name)
    sheet = wb.active
    driver.get(company_links[i])
    url = driver.current_url
    sheet.cell(row=2 + i, column=5).value= url
    driver.get('http://www.alexa.com/siteinfo/' + url)
    if check_exists_by_xpath('//*[@id="traffic-rank-content"]/div/span[2]/div[1]/span/span/div/strong'):
        rating = driver.find_element_by_xpath('//*[@id="traffic-rank-content"]/div/span[2]/div[1]/span/span/div/strong').get_attribute('innerHTML')
        rating = rating.split("\n")[2]
        rating = rating.replace(',','')
        sheet.cell(row=2 + i, column=1).value = rating
        if rating == '<span style="margin-left: 10px;">-</span>              ':
            sheet.cell(row=2 + i, column=1).value = 'No rating'
    else:
        sheet.cell(row=2 + i, column=1).value = 'No rating'

    wb.save(file_name)
    print('progress = '+ str(i) + '/' +str(len(company_links)))

    # sleep(randint(1,3))


