import requests
import time
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from xlsxwriter import Workbook
import xlrd
import re
import json
from openpyxl import load_workbook


workbook = xlrd.open_workbook('forex_10emp_900.xlsx')
worksheet = workbook.sheet_by_index(0)
rows = worksheet.nrows

links = []
for i in range (0,rows - 1):
    links.append(worksheet.cell_value(1 + i,2))


ua = UserAgent()
header = {'user-agent': ua.chrome}
company_links = []

for i in range(0,len(links)):

    while True:
        ua = UserAgent()
        header = {'user-agent': ua.chrome}
        page = requests.get(
            links[i],
            headers=header)
        soup = BeautifulSoup(page.content, 'lxml')
        # print(page.status_code)
        a = soup.find_all('code', {"id" : "stream-footer-embed-id-content" })

        # print(a.__len__())

        # b= a.__len__()
        if len(a) == 0:
            if len(soup.find_all('code')) > 0 :
                print(soup.find_all('code'))
            pass
            # print('!!!!!!!')
            # print(page.cookies)
            # f = open('out.html', 'w')
            # f.write(soup.prettify())
            # f.close()
        else:
            break


    a = soup.find_all('code', {"id" : "stream-footer-embed-id-content" })

    m = re.search("<!\-\-.+\-\->", str(a))


    if m:
         parsedStr = m.group(0)[4:-3]
         parsedJson = json.loads(parsedStr)
         jsonKeys = parsedJson.keys()
         if 'website' in jsonKeys:
             company_links.append(parsedJson['website'])
             print('progress=',len(company_links),'/',len(links))
         else:
             company_links.append('No site')
    else:
        company_links.append('No site')

for el in company_links:
    print(el)
wb = load_workbook(filename='forex_10emp_900.xlsx')
sheet = wb.active
for i in range(0,len(company_links)):
    sheet.cell(row=2 + i, column=4).value = company_links[i]
wb.save('forex_10emp_900.xlsx')




